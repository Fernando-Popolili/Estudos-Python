from openpyxl import load_workbook

# --- PARTE 1: ABRIR PLANILHA DE VENDAS ---

caminho_vendas = "C:/Users/PICHAU/Documents/github/AUTOMAÇÕES/vendas_de_lanches.xlsx"
caminho_txt = "C:/Users/PICHAU/Documents/github/AUTOMAÇÕES/anotações.txt"

# Carregar workbook existente
planilha_vendas = load_workbook(caminho_vendas)

# Selecionar aba
aba_vendas = planilha_vendas['Sheet']

print("\n=== Dados atuais da planilha ===")
for linha in aba_vendas.iter_rows(values_only=True):
    print(linha)


# --- PARTE 2: ADICIONAR INFORMAÇÕES DO TXT NA ABA "Sheet" ---

with open(caminho_txt, "r", encoding="utf-8") as arquivo:
    for linha in arquivo:
        linha = linha.strip()               # remove \n
        valores = linha.split(",")          # divide por vírgula
        aba_vendas.append(valores)          # adiciona no final da aba


# --- PARTE 3: SALVAR A PLANILHA MODIFICADA ---

planilha_vendas.save(caminho_vendas)

print("\nDados do TXT adicionados com sucesso!")
