from openpyxl import load_workbook, Workbook

# --- PARTE 1: LER EXCEL ---

# Abrindo excel com caminho
planilha_vendas = load_workbook('C:/Users/PICHAU/Documents/github/AUTOMAÇÕES/vendas_de_lanches.xlsx')

# Abrindo a aba correta
aba_vendas = planilha_vendas['Sheet']

print("\n=== Dados da planilha de vendas ===")
for linha in aba_vendas.iter_rows(values_only=True):
    print(linha)


# --- PARTE 2: CRIAR NOVA PLANILHA COM TXT ---

planilha_contas = Workbook()
pagina1 = planilha_contas.active

with open("anotações.txt", "r", encoding="utf-8") as arquivo:
    for linha in arquivo:
        # remove quebras de linha
        linha = linha.strip()
        # transforma em lista separada por vírgulas
        pagina1.append(linha.split(","))

planilha_contas.save("contas_a_pagar.xlsx")

print("\nArquivo 'contas_a_pagar.xlsx' criado com sucesso!")
